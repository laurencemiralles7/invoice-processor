"""
Load the margin sheet and check each invoice row against it.
Uses the publicly accessible Google Sheet export (no auth needed for reads).
"""
import urllib.request
import io
import openpyxl
from config import (
    MARGIN_SHEET_ID,
    MARGIN_COL_TITLE, MARGIN_COL_VARIANT, MARGIN_COL_SKU,
    MARGIN_COL_PRICE, MARGIN_COL_BUYING_USD,
    MARGIN_GREEN_MAX, MARGIN_ORANGE_MAX,
)

# Cache loaded margin tabs in memory so we don't re-fetch per row
_margin_cache = {}


def _fetch_margin_sheet():
    """Download the full margin sheet XLSX once and cache it."""
    global _margin_cache
    if _margin_cache:
        return _margin_cache

    url = f"https://docs.google.com/spreadsheets/d/{MARGIN_SHEET_ID}/export?format=xlsx"
    print("  Downloading margin sheet...")
    with urllib.request.urlopen(url) as resp:
        data = resp.read()

    wb = openpyxl.load_workbook(io.BytesIO(data))
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        products = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # skip header
            sku = row[MARGIN_COL_SKU]
            title = row[MARGIN_COL_TITLE]
            price = row[MARGIN_COL_PRICE]
            buying = row[MARGIN_COL_BUYING_USD]
            variant = row[MARGIN_COL_VARIANT]

            if not price or not buying:
                continue
            try:
                price = float(price)
                buying = float(buying)
            except (TypeError, ValueError):
                continue

            be_roas = price / (price - buying) if (price - buying) > 0 else None
            cogs_pct = (buying / price * 100) if price > 0 else None

            products.append({
                "sku": str(sku).strip() if sku else None,
                "sku_base": _base_sku(sku),
                "title": str(title).strip() if title else None,
                "title_lower": str(title).strip().lower() if title else None,
                "variant": str(variant).strip() if variant else None,
                "price": price,
                "buying_usd": buying,
                "be_roas": be_roas,
                "cogs_pct": cogs_pct,
            })
        _margin_cache[sheet_name] = products

    return _margin_cache


def _base_sku(sku):
    if not sku:
        return None
    s = str(sku).strip()
    parts = s.split("-")
    if len(parts) >= 2:
        return "-".join(parts[:2])
    return s


def _margin_color(be_roas):
    if be_roas is None:
        return "UNKNOWN"
    if be_roas < MARGIN_GREEN_MAX:
        return "GREEN"
    if be_roas <= MARGIN_ORANGE_MAX:
        return "ORANGE"
    return "RED"


def _find_product(products, erp_sku, variant_sku, title):
    """
    Try to find a margin sheet product by SKU match (several strategies), then title.

    Match order:
      1. Exact ERP SKU
      2. ERP base SKU (FRUGAZ-0673 matches FRUGAZ-0673-19)
      3. ERP prefix: invoice starts with margin_sku (e.g. DSHD3-BK2020 → BK202015)
      4. ERP prefix -1: invoice starts with margin_sku[:-1] (e.g. DSHD5-XM8381 → XM838XXX)
      5. Same four strategies for variant SKU
      6. Title (case-insensitive, exact)
    """
    erp_base = _base_sku(erp_sku)
    var_base = _base_sku(variant_sku)
    erp_str = str(erp_sku).strip() if erp_sku else None
    var_str = str(variant_sku).strip() if variant_sku else None
    MIN_PREFIX = 8  # safety: only fuzzy-match if prefix is at least this long

    for candidate_sku in [erp_str, erp_base, var_str, var_base]:
        if not candidate_sku:
            continue
        for p in products:
            m = p["sku"]
            if not m:
                continue
            # 1. exact
            if m == candidate_sku:
                return p
            # 2. base match
            if p["sku_base"] == _base_sku(candidate_sku):
                return p
            # 3. invoice starts with margin SKU (margin is prefix of invoice)
            if len(m) >= MIN_PREFIX and candidate_sku.startswith(m):
                return p
            # 4. invoice starts with margin SKU minus last char (family prefix)
            family = m[:-1]
            if len(family) >= MIN_PREFIX and candidate_sku.startswith(family):
                return p

    # 5. Title match (case-insensitive exact)
    if title:
        title_lower = str(title).strip().lower()
        for p in products:
            if p["title_lower"] and p["title_lower"] == title_lower:
                return p

    return None


def check_margins(invoice_data, margin_tab):
    """
    For each row in invoice_data["rows"], look up the product in the margin sheet
    and attach margin info. Mutates rows in place.

    Adds to each row:
        margin_product: matched product dict or None
        be_roas: float or None
        cogs_pct: float or None
        margin_color: "GREEN" | "ORANGE" | "RED" | "UNKNOWN" | "NOT_FOUND"
        margin_note: short human-readable note
    """
    all_tabs = _fetch_margin_sheet()
    products = all_tabs.get(margin_tab, [])

    if not products:
        print(f"  Warning: margin tab '{margin_tab}' not found or empty.")

    not_found = []

    for row in invoice_data["rows"]:
        product = _find_product(
            products,
            row.get("erp_sku"),
            row.get("variant_sku"),
            row.get("title"),
        )

        if product is None:
            row["margin_product"] = None
            row["be_roas"] = None
            row["cogs_pct"] = None
            row["margin_color"] = "NOT_FOUND"
            row["margin_note"] = "Not in margin sheet — add manually"
            not_found.append(row)
        else:
            color = _margin_color(product["be_roas"])
            row["margin_product"] = product
            row["be_roas"] = product["be_roas"]
            row["cogs_pct"] = product["cogs_pct"]
            row["margin_color"] = color
            row["margin_note"] = _note(color, product)

    invoice_data["margin_not_found"] = not_found
    return invoice_data


def _note(color, product):
    roas = f"{product['be_roas']:.2f}" if product["be_roas"] else "N/A"
    cogs = f"{product['cogs_pct']:.1f}%" if product["cogs_pct"] else "N/A"
    if color == "GREEN":
        return f"OK — BE-ROAS {roas}, COGs {cogs}"
    if color == "ORANGE":
        return f"Monitor — BE-ROAS {roas}, COGs {cogs}"
    if color == "RED":
        return f"ACTION — BE-ROAS {roas}, COGs {cogs}"
    return f"BE-ROAS {roas}, COGs {cogs}"
