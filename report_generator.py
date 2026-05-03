"""
Generate outputs from processed invoice data:
  1. Styled XLSX with color-coding, daily subtotals
  2. Cassy Slack message text
  3. P&L update preview (date → COG amount)
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from collections import defaultdict
from config import STORES


# ── Colors ────────────────────────────────────────────────────────────────────
FILL_RED       = PatternFill("solid", fgColor="FFCCCC")   # refund
FILL_ORANGE    = PatternFill("solid", fgColor="FFE0B2")   # orange margin
FILL_GREEN     = PatternFill("solid", fgColor="C8E6C9")   # green margin (light, optional)
FILL_YELLOW    = PatternFill("solid", fgColor="FFFF99")   # not found in margin sheet
FILL_SUBTOTAL  = PatternFill("solid", fgColor="D9D9D9")   # daily subtotal row
FONT_BOLD      = Font(bold=True)
FONT_RED       = Font(bold=True, color="CC0000")
THIN           = Side(style="thin")
BORDER_ALL     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cell_fill(ws, row, col, fill):
    ws.cell(row=row, column=col).fill = fill


def _apply_row_fill(ws, row_num, num_cols, fill):
    for c in range(1, num_cols + 1):
        ws.cell(row=row_num, column=c).fill = fill


def generate_xlsx(invoice_data, output_path):
    """Write a styled, color-coded invoice XLSX with daily subtotal rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    headers = [
        "Order Name", "Paid At", "Customer", "Country",
        "ERP SKU", "Title", "Qty", "Unit Price", "Discount", "In Total",
        "Margin Status", "BE-ROAS", "COGs%", "Note",
    ]
    num_cols = len(headers)

    # Header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = FONT_BOLD
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Group rows by date for subtotals
    rows_by_date = defaultdict(list)
    for r in invoice_data["rows"]:
        key = r["paid_at"] or date(1970, 1, 1)
        rows_by_date[key].append(r)

    current_row = 2

    for day in sorted(rows_by_date.keys()):
        day_rows = rows_by_date[day]
        day_total = sum(r["in_total"] for r in day_rows)

        for r in day_rows:
            color = r.get("margin_color", "UNKNOWN")

            values = [
                r["order_name"],
                str(r["paid_at"]) if r["paid_at"] else "",
                r["customer"],
                r["country"],
                r["erp_sku"],
                r["title"],
                r["quantity"],
                r["unit_price"],
                r["discount"],
                r["in_total"],
                color,
                round(r["be_roas"], 2) if r["be_roas"] else "",
                round(r["cogs_pct"], 1) if r["cogs_pct"] else "",
                r.get("margin_note", ""),
            ]

            for c, v in enumerate(values, 1):
                ws.cell(row=current_row, column=c, value=v)

            # Row fill based on status
            if r["is_refund"]:
                _apply_row_fill(ws, current_row, num_cols, FILL_RED)
                ws.cell(row=current_row, column=1).font = FONT_RED
            elif color == "RED":
                _apply_row_fill(ws, current_row, num_cols, FILL_ORANGE)
            elif color == "NOT_FOUND":
                _apply_row_fill(ws, current_row, num_cols, FILL_YELLOW)
            elif color == "GREEN":
                pass  # leave white for readability

            current_row += 1

        # Subtotal row for this date
        ws.cell(row=current_row, column=1, value=f"SUBTOTAL — {day}")
        ws.cell(row=current_row, column=10, value=round(day_total, 2))
        ws.cell(row=current_row, column=1).font = FONT_BOLD
        ws.cell(row=current_row, column=10).font = FONT_BOLD
        _apply_row_fill(ws, current_row, num_cols, FILL_SUBTOTAL)
        current_row += 1

    # Grand total row
    current_row += 1
    ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    ws.cell(row=current_row, column=10, value=round(invoice_data["grand_total"], 2))
    ws.cell(row=current_row, column=1).font = FONT_BOLD
    ws.cell(row=current_row, column=10).font = FONT_BOLD

    # Column widths
    widths = [14, 12, 22, 18, 20, 38, 5, 11, 10, 10, 14, 9, 8, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Issues summary tab
    issues = [r for r in invoice_data["rows"]
              if r.get("margin_color") in ("RED", "NOT_FOUND") or r["is_refund"]]
    if issues:
        ws2 = wb.create_sheet("Issues")
        ws2.cell(row=1, column=1, value="Issues to Action").font = FONT_BOLD
        ws2.cell(row=2, column=1, value="Type")
        ws2.cell(row=2, column=2, value="Order")
        ws2.cell(row=2, column=3, value="SKU")
        ws2.cell(row=2, column=4, value="Title")
        ws2.cell(row=2, column=5, value="In Total")
        ws2.cell(row=2, column=6, value="Note")
        for c in range(1, 7):
            ws2.cell(row=2, column=c).font = FONT_BOLD

        r_num = 3
        for r in issues:
            if r["is_refund"]:
                issue_type = "REFUND"
            elif r.get("margin_color") == "RED":
                issue_type = "RED MARGIN"
            else:
                issue_type = "NOT IN SHEET"
            ws2.cell(row=r_num, column=1, value=issue_type)
            ws2.cell(row=r_num, column=2, value=r["order_name"])
            ws2.cell(row=r_num, column=3, value=r["erp_sku"])
            ws2.cell(row=r_num, column=4, value=r["title"])
            ws2.cell(row=r_num, column=5, value=r["in_total"])
            ws2.cell(row=r_num, column=6, value=r.get("margin_note", ""))
            r_num += 1

        ws2.column_dimensions["D"].width = 38
        ws2.column_dimensions["F"].width = 40

    # output_path can be a file path string or a BytesIO buffer
    wb.save(output_path)
    if isinstance(output_path, str):
        print(f"  Saved: {output_path}")


def generate_cassy_message(invoice_data, store_name, invoice_filename):
    """Build the #hst-invoices / #dayone-invoices Cassy tag message."""
    daily = invoice_data["daily_totals"]
    grand = invoice_data["grand_total"]
    issues = invoice_data.get("margin_not_found", [])
    red_rows = [r for r in invoice_data["rows"] if r.get("margin_color") == "RED"]
    refunds = invoice_data["refund_rows"]

    lines = []
    lines.append(f"@Cassy — {store_name} invoice reviewed and verified.")
    lines.append("")
    lines.append(f"Invoice Breakdown ({invoice_filename}):")
    for d, total in sorted(daily.items()):
        lines.append(f"  {store_name} {d}: ${total:.2f}")
    lines.append(f"  Total: ${grand:.2f}")

    if refunds:
        lines.append("")
        lines.append(f"Refunds ({len(refunds)} entries highlighted red in invoice):")
        for r in refunds:
            lines.append(f"  - {r['order_name']} | {r['title']} | {r['customer']}")

    if red_rows:
        lines.append("")
        lines.append(f"Red Margin Items ({len(red_rows)} items — action required):")
        seen = set()
        for r in red_rows:
            key = r.get("erp_sku") or r.get("title")
            if key not in seen:
                seen.add(key)
                roas = f"{r['be_roas']:.2f}" if r["be_roas"] else "N/A"
                lines.append(f"  - {r['title']} (SKU: {r['erp_sku']}) — BE-ROAS {roas}")

    if issues:
        lines.append("")
        lines.append(f"Not in Margin Sheet ({len(issues)} products — please add manually):")
        seen = set()
        for r in issues:
            key = r.get("erp_sku") or r.get("title")
            if key not in seen:
                seen.add(key)
                lines.append(f"  - {r['title']} (SKU: {r['erp_sku']})")

    if not refunds and not red_rows and not issues:
        lines.append("")
        lines.append("No issues found. All margins verified.")

    return "\n".join(lines)


def generate_pl_summary(invoice_data, store_name):
    """Return a human-readable P&L update preview table."""
    lines = []
    lines.append(f"P&L Update — {store_name}")
    lines.append(f"{'Date':<14} {'COG':<10}")
    lines.append("-" * 26)
    for d, total in sorted(invoice_data["daily_totals"].items()):
        lines.append(f"{str(d):<14} ${total:.2f}")
    lines.append("-" * 26)
    lines.append(f"{'TOTAL':<14} ${invoice_data['grand_total']:.2f}")
    return "\n".join(lines)
